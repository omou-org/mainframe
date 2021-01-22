from django.shortcuts import render
from django.http import HttpResponse

from course.models import Course, CourseCategory

from rest_framework.views import APIView
from .serializers import UploadSerializer
# from rest_framework import MultiPartParser
from .onboarding import Upload

from openpyxl import load_workbook
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

from .process_files import process_import

class FileUpload(APIView):
    # parser_classes = [MultiPartParser, FormParser]

    def post(self, request, format=None):
        file_upload = request.FILES['file_upload']
        file_name = file_upload.name
        file_type = request.data['file_type']
        if file_name != file_choices[file_type]:
            return HttpResponse("wrong file type")

        path = default_storage.save('tmp/curr.xlsx', ContentFile(file_upload.read()))
        wb = load_workbook(filename=path)
        for row in wb.iter_rows(min_row=1, max_row=1):
            for i, cell in enumerate(row):
                if cell != column_names[file_type][i]:
                    return HttpResponse(
                        """
                        Columns are either named incorrectly or out of order, 
                        please look at the template downloaded on the previous page
                        """)
                
        print('process import' , process_import(path, file_type))

        default_storage.delete(path)
        return HttpResponse("success")


file_choices = {
    "student" : "students.xlsx",
    "parent" : "parents.xlsx",
    "course" : "course.xlsx",
    "instructor" : "instructors.xlsx",
    "course_categories" : "course_categories.xlsx"
}

column_names = {
    "student": [
        'first name',
        'last name',
        'email',
        'date of birth',
        'school',
        'grade',
        'parent first name',
        'parent email'
    ],
    "parent" : [
        "first name",
        "last name",
        "email",
        "phone",
        "zip code"
    ],
    "instructor" : [
        "first name",
        "last name",
        "email",
        "phone",
        "date of birth",
        "teaching subjects",
        "biography",
        "years of experience"
    ],
    "course_categories": [
        "name",
        "description"
    ]
}