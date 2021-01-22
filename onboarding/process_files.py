from openpyxl import load_workbook
from account.models import Student, Parent, Instructor
from course.models import CourseCategory
from django.contrib.auth.models import User
import random, string

def process_course_categories(sheet):
    columns = {
        "name": 0,
        "description": 1
    }
    for i, row in enumerate(sheet.iter_rows(min_row=3)):
        info = [i for i in row]
        new_course_category = CourseCategory(
            name=info[columns["name"]],
            description=info[columns["description"]]
            )
        new_course_category.save()


def process_instructors(sheet):
    columns = {
        "first name": 0,
        "last name": 1,
        "email": 2,
        "phone": 3,
        "category": 6,
        "biography": 7,
        "years experience": 8
    }
    last_col = sheet.max_column
    print("last col", last_col - 1)
    for i, row in enumerate(sheet.iter_rows(min_row=3)):
        info = [i for i in row]
        print("info", info)
        username = str(random.randint(1, 1000000))
        new_user = User(
            username=username,
            password=''.join(random.choice(string.ascii_letters) for _ in range(10)),
            first_name=info[columns["first name"]].value,
            last_name=info[columns["last name"]].value,
            email = info[columns["email"]].value,
        )
        new_user.save()

        new_instructor = Instructor(
            user=new_user,
            phone_number=info[columns["phone"]].value,
            biography=info[columns["biography"]].value,
            experience=info[columns["years experience"]].value
            
        )
        categories: str = info[6]
        category_list = categories.split(",")
        for category in category_list:
            if category not in [course.name for course in CourseCategory.objects.all()]:
                c = CourseCategory(name=category)
                c.save()
            else:
                existing_categories = CourseCategory.objects.filter(subjects__in=[category_list])
                for subject in existing_categories:
                    new_instructor.subjects.add(subject)
                    
        new_instructor.save()

def process_students(sheet):
    columns = {
        "first name": 0,
        "last name": 1,
        "address": 2,
        "city": 3,
        "state": 4,
        "zipcode": 5,
        "grade": 6,
        "phone number": 7,
        "school": 8,
        "birth date": 9,
        "parent": 10
    }
    for i, row in enumerate(sheet.iter_rows(min_row=3)):
        info = [i for i in row]
        username = str(random.randint(1, 1000000))
        new_user = User(
            username=username,
            password=''.join(random.choice(string.ascii_letters) for _ in range(10)),
            first_name=info[columns["first name"]].value,
            last_name=info[columns["last name"]].value,
        )
        new_user.save()
        
        new_student = Student(
            user=new_user,
            phone_number=info[columns["phone number"]].value,
            grade=info[columns["grade"]].value,
            school=info[columns["school"]].value,
            birth_date=info[columns["birth date"]].value,
            address=info[columns["address"]].value,
            city=info[columns["city"]].value,
            state=info[columns["state"]].value,
            zipcode=info[columns["zipcode"]].value
        )
        new_student.save()

def process_parents(sheet):
    columns = {
        "first name": 0,
        "last name": 1,
        "address": 2,
        "city": 3,
        "state": 4,
        "zipcode": 5,
        "phone number": 6,
        "relationship": 7
    }
    for i, row in enumerate(sheet.iter_rows(min_row=3)):
        info = [i for i in row]
        username = str(random.randint(1, 1000000))

        new_user = User(
            username=username,
            password=''.join(random.choice(string.ascii_letters) for _ in range(10)),
            first_name=info[columns["first name"]].value,
            last_name=info[columns["last name"]].value,
        )
        new_user.save()

        new_parent = Parent(
            user=new_user,
            phone_number=info[columns["phone number"]].value,
            address=info[columns["address"]].value,
            city=info[columns["city"]].value,
            state=info[columns["state"]].value,
            zipcode=info[columns["zipcode"]].value,
            relationship=info[columns["relationship"]].value
        )
        new_parent.save()

def process_import(path, file_type):
    file_switcher = {
        "student": process_students,
        "instructor": process_instructors,
        "parent": process_parents,
        "course_categories": process_course_categories,
        }

    wb = load_workbook(filename=path)
    sheet = wb.active
    file_switcher.get(file_type)(sheet)
    return (path, file_type)



column_names = {
    "student": [
        'first name',
        'last name',
        'email',
        'date of birth',
        'school',
        'grade',
        'parent first name',
        'parent email'
    ],
    "parent" : [
        "first name",
        "last name",
        "email",
        "phone",
        "zip code"
    ],
    "instructor" : [
        "first name",
        "last name",
        "email",
        "phone",
        "date of birth",
        "teaching subjects",
        "biography",
        "years of experience"
    ],
    "course_categories": [
        "name",
        "description"
    ]
}