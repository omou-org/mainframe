from graphene import Field, ID, List, String
from graphene_django.types import DjangoObjectType
from graphql_jwt.decorators import login_required, staff_member_required

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from tempfile import NamedTemporaryFile
import base64

from mainframe.permissions import IsOwner
from django_graphene_permissions import permissions_checker

from account.models import (
    Admin,
    Student,
    Parent,
    Instructor
)
from course.models import Course
from onboarding.models import Business, BusinessAvailability


class BusinessAvailabilityType(DjangoObjectType):
    class Meta:
        model = BusinessAvailability


class BusinessType(DjangoObjectType):
    availability_list = List(BusinessAvailabilityType, source='availability_list')

    class Meta:
        model = Business


def autosize_ws_columns(worksheet):
    for col in worksheet.columns:
        column = get_column_letter(col[0].column)
        max_length = max(len(str(cell.value)) for cell in col[1:])
        adjusted_width = (max_length+2)*1.1
        worksheet.column_dimensions[column].width = adjusted_width


def style_instruction_comment(cell):
    cell.alignment = Alignment(wrapText=True)
    cell.font = Font(color='44b6d9')


def workbook_to_base64(wb):
    with NamedTemporaryFile() as tmp:
        wb.save(filename = tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    base64_stream = base64.b64encode(stream).decode("utf-8")
    return base64_stream


def create_accounts_template(show_errors=False):
    wb = Workbook()
    wb.create_sheet("Parents")
    wb.create_sheet("Students")
    wb.create_sheet("Instructors")

    # parent sheet
    parents_ws = wb.get_sheet_by_name("Parents")
    parents_ws["A1"].value = "Instructions: Please enter parent details in this spreadsheet similar to the example on row 3. Double check that all emails, phone numbers, and zip codes are valid, otherwise those parents may not be uploaded correctly. The first row is an example that will not be uploaded."
    parents_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    style_instruction_comment(parents_ws["A1"])

    parents_column_names = ["First Name", "Last Name", "Email", "Phone", "Zip Code (Optional)"]
    parents_example = ["Ken", "Chan", "kennyken@yahoo.com",	"5635573354", "94345"]
    if show_errors:
        parents_column_names.insert(0, "Error Message")
        parents_example.insert(0, "")
    
    for c in range(len(parents_column_names)):
        parents_ws.cell(row=2, column=c+1).value = parents_column_names[c]
        parents_ws.cell(row=3, column=c+1).value = parents_example[c]

    # students
    students_ws = wb.get_sheet_by_name("Students")
    students_ws["A1"].value = "Instructions: Please enter student details in this spreadsheet similar to the example on row 3. Please enter the student's parent's first last name and email exactly as was entered in the parent spreadsheet, otherwise the student will not be visible under the parent's profile in Omou. The first row is an example that will not be uploaded."
    students_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    style_instruction_comment(students_ws["A1"])

    students_column_names = ["First Name", "Last Name", "Email", "Birthday MM/DD/YYYY (Optional)", "School (Optional)", "Grade Level (Optional)", "Parent's First Name", "Parent's Last Name", "Parent's Email"]
    students_example = ["Kevin", "Chan", "kev.chan@gmail.com", "08/08/2008", "Alamador High School", "11", "Ken", "Chan", "kennyken@yahoo.com"]
    if show_errors:
        students_column_names.insert(0, "Error Message")
        students_example.insert(0, "")

    for c in range(len(students_column_names)):
        students_ws.cell(row=2, column=c+1).value = students_column_names[c]
        students_ws.cell(row=3, column=c+1).value = students_example[c]

    # instructors
    instructors_ws = wb.get_sheet_by_name("Instructors")
    instructors_ws["A1"].value = "Instructions: Please enter instructor details in this spreadsheet similar to the example on row 3. Instructors with invalid emails will not be entered into Omou, so please confirm the instructor's email is valid prior to submission! The first row is an example that will not be uploaded."
    instructors_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    style_instruction_comment(instructors_ws["A1"])

    instructors_column_names = ["First Name", "Last Name", "Email", "Phone", "Biography (Optional)", "Years of Experience (Optional)", "Address (Optional)", "City", "State", "Zip Code"]
    instructors_example = ["Vicky", "Lane", "vlane@gmail.com", "6453456765", "Vicky is experienced in teaching math for all skill levels.", "5", "456 Candy Lane", "Los Angeles", "CA", "90034"]
    if show_errors:
        instructors_column_names.insert(0, "Error Message")
        instructors_example.insert(0, "")

    for c in range(len(instructors_column_names)):
        instructors_ws.cell(row=2, column=c+1).value = instructors_column_names[c]
        instructors_ws.cell(row=3, column=c+1).value = instructors_example[c]

    # remove default sheet
    del wb['Sheet']

    # formatting
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        if sheet.sheet_state == 'hidden':
            continue
        autosize_ws_columns(sheet)
        # bold column headers
        for cell in sheet["2:2"]:
            cell.font = Font(bold=True)
        # increase comment row height
        sheet.row_dimensions[1].height=70
        # freeze top 2 rows
        sheet.freeze_panes='A3'
    
    return wb


# business_id populates instructors from same business
# show_errors=True adds error message column
def create_course_templates(business_id, show_errors=False):
    wb = Workbook()
    wb.create_sheet("Instructor Roster (hidden)")
    wb.create_sheet("Step 1 - Subject Categories")
    wb.create_sheet("Step 2 - Classes")

    # instructors populate dropdown
    instructors_ws = wb.get_sheet_by_name("Instructor Roster (hidden)")
    instructors_ws.sheet_state = 'hidden'
    instructors_ws.cell(row=1, column=1).value = "Instructor Name (email)"
    total_instructors = Instructor.objects.business(business_id).count()
    for row, instructor in enumerate(Instructor.objects.business(business_id)):
        combined_value = f"{instructor.user.first_name} {instructor.user.last_name} ({instructor.user.email})"
        instructors_ws.cell(row=row+2, column=1).value = combined_value

    # subject categories
    categories_ws = wb.get_sheet_by_name("Step 1 - Subject Categories")

    categories_ws.cell(row=1, column=1).value = "Instructions: Please FIRST fill out the list of subject or course category in the first column. This will become a list of subjects to fill out the required course subject field in Step 2 - Classes sheet. The first row is an example that will not be uploaded."
    categories_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    style_instruction_comment(categories_ws["A1"])

    categories_column_names = ["Subjects", "Description"]
    categories_example = ["SAT Prep", "Preparational courses for students taking the SAT"]
    if show_errors:
        categories_column_names.insert(0, "Error Message")
        categories_example.insert(0, "")

    for c in range(len(categories_column_names)):
        categories_ws.cell(row=2, column=c+1).value = categories_column_names[c]
        categories_ws.cell(row=3, column=c+1).value = categories_example[c]

    # classes
    course_ws = wb.get_sheet_by_name("Step 2 - Classes")

    course_ws.cell(row=1, column=1).value = "Instructions: Please fill out the course details similar to the example in row 3. You may only select subjects previous defined in Step 1 - Subject Categories. You may only select instructors that are already in the database. The first row is an example that will not be uploaded."    
    course_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    style_instruction_comment(course_ws["A1"])

    course_ws.cell(row=1, column=10).value = 'Instructions: 5 course session slots are provided in format of "Session Day" "Start Time" and "End Time." "Session Day" is the day of week the session takes place on. If a course meets more than 5 times a week, please add additional "Session Day" "Start Time" and "End Time" columns in increasing order (e.g. "Session Day 6"). If a course only meets once, you may leave the other session slots empty.'
    course_ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=14)
    style_instruction_comment(course_ws["J1"])

    class_column_names = ["Course Name", "Instructor", "Instructor Confirmed? (Y/N)", "Subject", "Course Description", "Academic Level", "Room Location", "Total Tuition", "Enrollment Capacity (>=4)", "Start Date", "End Date", "Session Day 1", "Start Time 1", "End Time 1", "Session Day 2", "Start Time 2", "End Time 2", "Session Day 3", "Start Time 3", "End Time 3", "Session Day 4", "Start Time 4", "End Time 4", "Session Day 5", "Start Time 5", "End Time 5"]
    class_example = ["SAT Math Prep", "Daniel Wong (email)", "Y", "SAT Prep", "This is a prep course for SAT math. Open for all grade levels.", "High School", "Online", "450.0", "5", "12/1/2021", "12/30/2021", "Wednesday", "4:00 PM",	"5:00 PM", "Friday", "2:00 PM", "3:00 PM"]
    if show_errors:
        class_column_names.insert(0, "Error Message")
        class_example.insert(0, "")

    """
    Validations:
    openpyxl bug: need to set showDropDown=False to show dropdown menu
    1048576 is the last row in excel
    """

    # date validation
    date_dv = DataValidation(
        type="date",
        allow_blank=False
        )
    course_ws.add_data_validation(date_dv)

    # day of week validation
    day_of_week_dv = DataValidation(
        type="list",
        formula1='"Monday,Tuesday,Wednesday,Thursday,Friday,Saturday,Sunday"',
        allow_blank=False,
        showDropDown=False
        )
    course_ws.add_data_validation(day_of_week_dv)

    # time validation
    time_dv = DataValidation(
        type="time",
        allow_blank=False,
        )
    course_ws.add_data_validation(time_dv)

    # academic level validation
    academic_level_dv = DataValidation(
        type="list",
        formula1='"Elementary,Middle School,High School,College"',
        allow_blank=False,
        showDropDown=False
        )
    course_ws.add_data_validation(academic_level_dv)

    # instructor validation
    if total_instructors > 0:
        instructor_dv = DataValidation(
            type="list",
            formula1='{0}!$A$2:$A${1}'.format(
                quote_sheetname("Instructor Roster (hidden)"),
                total_instructors
                ),
            allow_blank=False,
            showDropDown=False
            )
        course_ws.add_data_validation(instructor_dv)

    # Y/N validation
    yes_no_dv = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=False,
        showDropDown=False
        )
    course_ws.add_data_validation(yes_no_dv)

    # subject validation
    subject_formula1 = '{0}!$B$4:$B$1048576' if show_errors else '{0}!$A$4:$A$1048576'
    subject_dv = DataValidation(
        type="list",
        formula1=subject_formula1.format(
            quote_sheetname("Step 1 - Subject Categories")
            ),
        allow_blank=False,
        showDropDown=False
        )
    course_ws.add_data_validation(subject_dv)

    # tuition validation
    tuition_dv = DataValidation(
        type="decimal",
        allow_blank=False,
        operator="greaterThan",
        formula1=0
    )
    course_ws.add_data_validation(tuition_dv)  
    
    # capacity validation
    capacity_dv = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=4
    )
    course_ws.add_data_validation(capacity_dv)


    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    # populate validations
    for c in range(len(class_column_names)):
        col = colnum_string(c+1)

        if class_column_names[c] == "Academic Level":
            academic_level_dv.add(f"{col}4:{col}1048576")
        
        if class_column_names[c] == "Instructor" and total_instructors > 0:
            instructor_dv.add(f"{col}4:{col}1048576")
        
        if class_column_names[c] == "Instructor Confirmed? (Y/N)":
            yes_no_dv.add(f"{col}4:{col}1048576")

        if class_column_names[c] == "Subject":
            subject_dv.add(f"{col}4:{col}1048576")

        if class_column_names[c] == "Total Tuition":
            tuition_dv.add(f"{col}4:{col}1048576")

        if class_column_names[c] == "Enrollment Capacity (>=4)":
            capacity_dv.add(f"{col}4:{col}1048576")

        if class_column_names[c] in ["Start Date", "End Date"]:
            date_dv.add(f"{col}4:{col}1048576")

        if class_column_names[c].startswith("Session Day"):
            day_of_week_dv.add(f"{col}4:{col}1048576")
            
        if class_column_names[c].startswith("Start Time") or class_column_names[c].startswith("End Time"):
            time_dv.add(f"{col}4:{col}1048576")

        course_ws.cell(row=2, column=c+1).value = class_column_names[c]
    
    for c in range(len(class_example)):
        course_ws.cell(row=3, column=c+1).value = class_example[c]

    # remove default sheet
    del wb['Sheet']

    # formatting
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        if sheet.sheet_state == 'hidden':
            continue
        autosize_ws_columns(sheet)
        # bold column headers
        for cell in sheet["2:2"]:
            cell.font = Font(bold=True)
        # increase comment row height
        sheet.row_dimensions[1].height=70
        # freeze top 2 rows
        sheet.freeze_panes='A3'
    
    return wb


# business_id used to populate students and courses from same business
# show_errors=True adds error message column

def create_enrollment_templates(business_id, show_errors=False):
    
    wb = Workbook()
    wb.create_sheet("Student Roster (hidden)")

    # students populate dropdown
    students_ws = wb.get_sheet_by_name("Student Roster (hidden)")
    students_ws.sheet_state = 'hidden'
    students_ws.cell(row=1, column=1).value = "Student Name (email)"
    total_students = Student.objects.business(business_id).count()
    for row, student in enumerate(Student.objects.business(business_id)):
        students_ws.cell(row=row+2, column=1).value = f"{student.user.first_name} {student.user.last_name} ({student.user.email})" 
        

    # course templates
    for course in Course.objects.business(business_id):
        sheet_name = f"{course.title} - {course.id}"
        wb.create_sheet(sheet_name)
        course_ws = wb.get_sheet_by_name(sheet_name)

        course_ws["A1"].value = "Instructions: This course enrollment sheet and course details are generated from saved courses in the Omou database. Row 9 and beyond are dropdowns of students existing in the Omou database. Please select a student from the dropdown list starting from row 9 to create an enrollment for this course."
        course_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        style_instruction_comment(course_ws["A1"])

        # course info
        def date_format(date):
            return f"{date.month}/{date.day}/{date.year}"

      
        course_info = {
            "Course Name": course.title,
            "Course ID": course.id,
            "Instructor": f"{course.instructor.user.first_name} {course.instructor.user.last_name}",
            "Subject": course.course_category.name,
            "Course Date": f"{date_format(course.start_date)} - {date_format(course.end_date)}",
            "Enrollment Capacity": course.max_capacity,
        }

        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
            )
        background_color = PatternFill(start_color="EBFAFF", end_color="EBFAFF", fill_type = "solid")

        for row, info_label in enumerate(course_info):
            info_label_cell = course_ws.cell(row=row+2, column=1)
            info_value_cell = course_ws.cell(row=row+2, column=2)

            def style_course_info_cell(cell):
                cell.border = thin_border
                cell.fill = background_color
                cell.alignment = Alignment(horizontal='left')

            info_label_cell.value = info_label
            info_label_cell.font = Font(color='1F83A2', bold=True)
            style_course_info_cell(info_label_cell)

            info_value_cell.value = course_info[info_label]
            info_value_cell.font = Font(color='1F83A2')
            style_course_info_cell(info_value_cell)

        # student header
        if show_errors:
            course_ws['A8'].value = "Error Message"
            course_ws['A8'].font = Font(bold=True)
            course_ws['B8'].value = "Students Enrolled"
            course_ws['B8'].font = Font(bold=True)
        else:
            course_ws['A8'].value = "Students Enrolled"
            course_ws['A8'].font = Font(bold=True)

        # student validation
        if total_students > 0:
            student_dv = DataValidation(
                type="list",
                formula1='{0}!$A$2:$A${1}'.format(
                    quote_sheetname("Student Roster (hidden)"),
                    total_students
                    ),
                allow_blank=False,
                showDropDown=False
                )
            course_ws.add_data_validation(student_dv)
            student_dv.add("A9:A1048576")

    # remove default sheet
    del wb['Sheet']

    # formatting
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        if sheet.sheet_state == 'hidden':
            continue
        autosize_ws_columns(sheet)
        # bold column headers
        for cell in sheet["8:8"]:
            cell.font = Font(bold=True)
        # increase comment row height
        sheet.row_dimensions[1].height=70
        # freeze top 2 rows
        sheet.freeze_panes='A9'
    
    return wb


class Query(object):
    business = Field(BusinessType)
    account_templates = String()
    course_templates = String()
    enrollment_templates = String()

    @login_required
    def resolve_business(self, info, **kwargs):
        user_id = info.context.user.id

        account_queries = [
            Student.objects.filter(user__id=user_id),
            Instructor.objects.filter(user__id=user_id),
            Parent.objects.filter(user__id=user_id),
            Admin.objects.filter(user__id=user_id),
        ]
        account = next(query.first() for query in account_queries if query.exists())

        return Business.objects.get(id=account.business.id)

    @login_required
    @permissions_checker([IsOwner])
    def resolve_account_templates(self, info, **kwargs):
        wb = create_accounts_template()
        return workbook_to_base64(wb)

    @login_required
    @permissions_checker([IsOwner])
    def resolve_course_templates(self, info, **kwargs):
        owner = Admin.objects.get(user=info.context.user)
        wb = create_course_templates(owner.business.id)
        return workbook_to_base64(wb)
    
    @login_required
    @permissions_checker([IsOwner])
    def resolve_enrollment_templates(self, info, **kwargs):
        owner = Admin.objects.get(user=info.context.user)
        wb = create_enrollment_templates(owner.business.id)
        return workbook_to_base64(wb)
